import openpyxl
from openpyxl import Workbook

class ParseExcel(object):
    '''处理excel文件'''

    def __init__(self, parent=None ,file_name=None):
        ''''''

        self.parent = parent
        self.book = openpyxl.load_workbook(file_name,data_only=True)
        self.sheet = self.book.active

        self.sheetTitle = self.sheet.cell(row=1,column=1).value
        self._nrows = self.sheet.max_row  # 文件总行数
        self.avaRows =  self._nrows - 2  # 文件有效行数
        self.avaCols = self.sheet.max_column

        # self.__headers = self.sheet.row_values(0)

        self.allHeaders = self.getRowValues(2)
        self.allUserData = []

        for i in range(1,self.avaRows+1):
            rowData = self.getRowValues(i+2)
            self.allUserData.append(rowData)

        print(self.allHeaders)


    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def iter_salary_line(self):
        for i in range(1,self.avaRows+1):
            rowData = self.getRowValues(i+2)
            row_info = zip(self.allHeaders,rowData)
            yield list(row_info)

    @property
    def headers(self):
        return self.allHeaders

