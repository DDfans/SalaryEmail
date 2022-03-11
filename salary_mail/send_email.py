from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr
from smtplib import SMTP_SSL, SMTP
from decimal import (Decimal)
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side
import os
import datetime


class SendEmail(object):

    def __init__(self, win, password, mail_title, mail_headers, user_data, excel_p):
        self.win = win
        self.__password = password
        self.mail_title = mail_title
        self.mail_headers = mail_headers
        self.all_user_data = user_data
        self.mail_content = self.win.mail_content
        self.sender_text = self.win.sender_text.get()  # 发件邮箱
        self.sender_name_text = self.win.sender_name_text.get()

        self.save_file_path = excel_p

        today = datetime.datetime.now()
        month = today.month
        if month == 1:
            month = 12
        else:
            month -= 1
        self.send_moth = month
        self.name_index = -1
        self.id_index = -1

        for i in range(0,len(self.mail_headers)):
            if self.mail_headers[i] == "姓名":
                self.name_index = i
            if self.mail_headers[i] == "员工号":
                self.id_index = i

    def send_email(self, index):
        if self.name_index == -1 or self.id_index == -1:
            tk.messagebox.showerror(title='错误', message='Excel表头缺少"姓名"或"员工号"')
            return
        try:
            row = self.all_user_data[index]
        except StopIteration:
            return
        rec_email = row[-1]

        if rec_email is None or len(rec_email) == 0:
            self.report_result(name=row[self.name_index],id=row[self.id_index],email=rec_email,flag=False)
            self.send_email(index=index + 1)
            return

        flag = True
        try:
            self._send_email(smtp=self.smtp, sender=self.sender_text, sender_name=self.sender_name_text, sign="fdd", date="2022",
                             info_row=row)
        except Exception as e:
            try:
                self._send_email(smtp=self.smtp, sender=self.sender_text, sender_name=self.sender_name_text, sign="fdd", date="2022",
                                 info_row=row)
            except Exception as e:
                flag = False
            else:
                pass
        else:
            pass

        self.report_result(name=row[self.name_index], id=row[self.id_index],email=row[-1], flag=flag)

        self.send_email(index=index+1)

    def report_result(self, name, id,email, flag):
        self.win.count_done_row()
        self.win.show_percent_run()
        self.win.result_list.insert('', 'end', values=(name, id,email, "成功！" if flag else "发送失败！！！"))

    def _send_email(self, smtp, sender, sender_name, sign, date, info_row):

        msg = self._make_mail_text(sender=sender, sender_name=sender_name, sign=sign, date=date,
                                   info_row=info_row)
        smtp.sendmail(from_addr=sender, to_addrs=[info_row[-1]], msg=msg)

    def _login_smpt(self):
        '''登陆邮箱'''
        try:
            sender_text = self.win.sender_text.get()
            smtp_server = self.win.smtp_text.get()
            port = int(self.win.port_text.get())

            password = self.__password
            port = int(port)
        except Exception as e:
            tk.messagebox.showerror(title='错误', message='数据库错误！\n{}'.format(e))
            return

        try:
            if port == 25:
                smtp = SMTP(host=smtp_server, port=port)
            elif port == 465:
                smtp = SMTP_SSL(host=smtp_server, port=port)
            else:
                raise ConnectionError('SMTP 端口错误')
            smtp.login(sender_text, password)
        except Exception as e:
            tk.messagebox.showerror(title='登陆错误', message='请检查账号信息是否正确！\n{}'.format(e))
            raise

        return smtp

    def _make_mail_text(self, sender, sender_name, sign, date, info_row):
        msg = MIMEMultipart()
        msg.attach(MIMEText(self.mail_content, 'plain', 'utf-8'))
        msg['From'] = formataddr([sender_name, sender])
        msg['To'] = formataddr([info_row[self.name_index],info_row[-1]])
        msg['Subject'] = self.win.subject.get()

        user_name = info_row[self.name_index]
        file_name = '{}{}月工资条.xlsx'.format(user_name, self.send_moth)
        file_path = self.save_file_path + '/' + file_name

        self._make_mail_excel(info_row, file_path)

        att = MIMEBase('application', 'octet-stream')  # 这两个参数不知道啥意思，二进制流文件
        att.set_payload(open(file_path, 'rb').read())
        # 此时的附件名称为****.xlsx，截取文件名
        att.add_header('Content-Disposition', 'attachment', filename=file_name)
        encoders.encode_base64(att)
        msg.attach(att)

        return msg.as_string()

    def _make_mail_excel(self, info_row,file_path):
        wb = Workbook()
        ws = wb.active

        for col in range(len(self.mail_headers)):
            title_cell = ws.cell(row=2, column=col + 1)
            title_cell.value = self.mail_headers[col]
            title_cell.font = Font(name='宋体', size=11, bold=False, italic=False, color='000000')
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            title_cell.border = Border(left=Side(border_style='thin', color='000000'),
                                       right=Side(border_style='thin', color='000000'),
                                       top=Side(border_style='thin', color='000000'),
                                       bottom=Side(border_style='thin', color='000000'))

            data_cell = ws.cell(row=3, column=col + 1)
            data_cell.value = info_row[col]
            data_cell.font = Font(name='宋体', size=11, bold=False, italic=False, color='000000')
            data_cell.alignment = Alignment(horizontal='center', vertical='center')
            data_cell.border = Border(left=Side(border_style='thin', color='000000'),
                                      right=Side(border_style='thin', color='000000'),
                                      top=Side(border_style='thin', color='000000'),
                                      bottom=Side(border_style='thin', color='000000'))

        head_cell = ws.cell(row=1, column=1)
        head_cell.value = self.mail_title
        head_cell.font = Font(name='宋体', size=14, bold=True, italic=False, color='000000')
        head_cell.alignment = Alignment(horizontal='center', vertical='center')
        head_cell.border = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))

        ws.row_dimensions[2].height = 70
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(self.mail_headers))


        wb.save(file_path)

    def run(self):
        try:
            self.smtp = self._login_smpt()  # 登陆邮箱
        except Exception as e:
            return

        self.send_email(index=0)
